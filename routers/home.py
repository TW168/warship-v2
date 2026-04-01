"""
routers/home.py — Home, Meeting Report, and Briefing page routes.

Includes:
  GET /                                      — Home page (weather images)
  GET /api/gas-prices                        — Latest gas prices from DB (JSON)
  GET /meeting-report                        — Meeting Report page (filter form)
  GET /api/meeting-report/results            — HTMX partial: query results as cards
  GET /briefing                              — VIP Operations Briefing sub-page
  GET /api/analytics/weight-by-year         — Monthly pick_weight per year series (JSON)
  GET /api/analytics/freight-lbs-by-year-mei   — Monthly lbs from frt_cost_breakdown_mei (JSON)
  GET /api/analytics/unit-frt-cost-john        — All rows from unit_frt_cost_john (JSON)
  GET /api/analytics/freight-cost-by-plant     — Annual YTD freight cost by plant from Excel (JSON)
  GET /api/analytics/product-trend-top         — Top N products by total weight (JSON)
  GET /api/analytics/product-trend-monthly     — Monthly trend for top products (JSON)
  GET /api/analytics/product-diversity         — Unique products per month (JSON)
"""

from collections import defaultdict
import datetime
from pathlib import Path
import re

import openpyxl
import pandas as pd
from fastapi import APIRouter, Query, Request
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import text

from database import connect_to_database
from utils.product_trend_service import (
    load_product_data_from_sp,
    get_top_products,
    get_monthly_trend,
    get_product_diversity_over_time,
)

router = APIRouter(tags=["Home"])
templates = Jinja2Templates(directory="templates")

# Create the DB engine once at module level (connection pool is reused per request)
_engine = connect_to_database()

# Meeting report SQL — groups customers into named locations, aggregates shipping metrics.
# Uses named parameters (:site, :product_group, :date) for safe parameterized execution.
_CONSIGNMENT_COUNT_SQL = text("""
    SELECT COUNT(*) AS consignment_count
    FROM (
        SELECT DISTINCT BL_Number, Carrier_ID, Truck_Appointment_Date
        FROM warship.vw_bl_lbs_cnt_carrier_customer
        WHERE site = :site
          AND product_group = :product_group
          AND Truck_Appointment_Date = :date
          AND Carrier_ID NOT IN ('SAIA-IP', 'CWF-IP')
          AND Ship_to_Customer IN (
              'AMTOPP WAREHOUSE - HOUSTON',
              'INTEPLAST GROUP CORP. (AMTOPP)',
              'INTEPLAST GROUP CORP.(AMTOPP ( CFP)',
              'PINNACLE FILMS'
          )
    ) AS sub
""")

_CUSTOM_COUNT_SQL = text("""
    SELECT COUNT(*) AS custom_count
)
)
async def product_trend_monthly(
    top_n: int = Query(default=5, ge=1, le=20, description="Number of top products")
) -> JSONResponse:
          AND Truck_Appointment_Date = :date
          AND Carrier_ID NOT IN ('SAIA-IP', 'CWF-IP')
          AND Ship_to_Customer NOT IN (
              'AMTOPP WAREHOUSE - HOUSTON',
              'INTEPLAST GROUP CORP. (AMTOPP)',
              'INTEPLAST GROUP CORP.(AMTOPP ( CFP)',
              'PINNACLE FILMS'
          )
    ) AS sub
""")

_MEETING_REPORT_SQL = text("""
    SELECT
        CASE
            WHEN Ship_to_Customer IN ('AMTOPP WAREHOUSE - HOUSTON')
                THEN 'Houston'
            WHEN Ship_to_Customer IN ('INTEPLAST GROUP CORP. (AMTOPP)')
                THEN 'Remington'
            WHEN Ship_to_Customer IN ('INTEPLAST GROUP CORP.(AMTOPP ( CFP)')
                THEN 'Phoenix'
            WHEN Ship_to_Customer IN ('PINNACLE FILMS')
                THEN 'Charlotte'
            ELSE 'Customers'
        END AS `Group`,

        SUM(pallet_count)                                       AS Pallets,
        SUM(pick_weight)                                        AS Weight,
        ROUND(SUM((Unit_Freight / 100.0) * Pick_Weight))        AS Freight,

        CASE
            WHEN SUM(pick_weight) > 0
                THEN ROUND(
                    SUM(Unit_Freight * Pick_Weight) / SUM(pick_weight),
                4)
            ELSE 0
        END AS Avg_Freight_per_lb

    FROM warship.vw_bl_lbs_cnt_carrier_customer
    WHERE site = :site
      AND product_group = :product_group
      AND Truck_Appointment_Date = :date
    GROUP BY
        CASE
            WHEN Ship_to_Customer IN ('AMTOPP WAREHOUSE - HOUSTON')
                THEN 'Houston'
            WHEN Ship_to_Customer IN ('INTEPLAST GROUP CORP. (AMTOPP)')
                THEN 'Remington'
            WHEN Ship_to_Customer IN ('INTEPLAST GROUP CORP.(AMTOPP ( CFP)')
                THEN 'Phoenix'
            WHEN Ship_to_Customer IN ('PINNACLE FILMS')
                THEN 'Charlotte'
            ELSE 'Customers'
        END
""")

_PALLET_ENTRY_EXIT_SQL = text("""
    SELECT
        SUM(avg_day_shift_in) + SUM(avg_night_shift_in) AS entry_to_date,
        SUM(avg_1st_shift_out)
          + SUM(avg_2nd_shift_out)
          + SUM(avg_3rd_shift_out) AS exit_to_date
    FROM warship.daily_shift_averages
    WHERE date BETWEEN :date_from AND :date_to
""")


_ASSETS_DIR = Path(__file__).parent.parent / "static" / "assets"
_NO_CACHE = {"Cache-Control": "no-store, no-cache, must-revalidate"}

_MTD_SITES = ("AMJK", "TXAS", "AMIN", "AMAZ")
_SITE_CONSIGNMENT_CUSTOMERS = {
    "AMJK": {
        "INTEPLAST GROUP CORP. (AMTOPP)",
        "INTEPLAST GROUP CORP.(AMTOPP ( CFP)",
        "PINNACLE FILMS",
        "AMTOPP WAREHOUSE - HOUSTON",
    },
    "AMIN": {
        "INTEPLAST GROUP CORP.(AMTOPP ( CFP)",
        "PINNACLE FILMS",
        "AMTOPP WAREHOUSE - HOUSTON",
        "INTEPLAST GROUP CORP.(AMTOPP)",
    },
    "AMAZ": {
        "INTEPLAST GROUP CORP. (AMTOPP)",
        "PINNACLE FILMS",
        "INTEPLAST GROUP CORP.(AMTOPP)",
    },
    "TXAS": {
        "INTEPLAST GROUP CORP. (AMTOPP)",
        "INTEPLAST GROUP CORP.(AMTOPP ( CFP)",
        "PINNACLE FILMS",
        "INTEPLAST GROUP CORP.(AMTOPP)",
    },
    "AMSC": {
        "INTEPLAST GROUP CORP. (AMTOPP)",
        "INTEPLAST GROUP CORP.(AMTOPP ( CFP)",
        "AMTOPP WAREHOUSE - HOUSTON",
        "INTEPLAST GROUP CORP.(AMTOPP)",
    },
}


def _normalize_customer_name(name: str) -> str:
    """Normalize customer names for robust consignment matching."""
    return re.sub(r"\s+", " ", str(name or "").upper()).strip()


def _resolve_month_window(year_month: str | None) -> tuple[datetime.date, datetime.date, datetime.date, str]:
    """Resolve selected month into month-start, month-end, effective MTD end, and label."""
    today = datetime.date.today()

    if year_month:
        try:
            selected_first = datetime.date.fromisoformat(f"{year_month}-01")
        except ValueError as exc:
            raise ValueError(f"Invalid year_month '{year_month}'. Expected YYYY-MM.") from exc
    else:
        selected_first = today.replace(day=1)

    if selected_first.month == 12:
        next_month_first = datetime.date(selected_first.year + 1, 1, 1)
    else:
        next_month_first = datetime.date(selected_first.year, selected_first.month + 1, 1)

    month_end = next_month_first - datetime.timedelta(days=1)
    effective_end = today if selected_first.year == today.year and selected_first.month == today.month else month_end
    month_label = selected_first.strftime("%Y/%m")
    return selected_first, month_end, effective_end, month_label


def _build_all_site_mtd_rows(
    conn,
    product_group: str,
    year_month: str | None = None,
) -> list[dict[str, object]]:
    """Build all-site MTD customer-vs-consignment totals from SP output using pandas."""
    month_start, _month_end, selected_date, month_label = _resolve_month_window(year_month)

    results: list[dict[str, object]] = []
    dbapi_conn = conn.connection

    for site_code in _MTD_SITES:
        cursor = dbapi_conn.cursor(dictionary=True)
        try:
            cursor.callproc(
                "sp_bl_lbs_cnt_carrier_customer",
                [
                    month_start.isoformat(),
                    selected_date.isoformat(),
                    site_code,
                    product_group,
                ],
            )

            sp_rows: list[dict[str, object]] = []
            for result_set in cursor.stored_results():
                for row in result_set.fetchall():
                    sp_rows.append(
                        {
                            "Ship_to_Customer": row.get("Ship_to_Customer"),
                            "pick_weight": row.get("pick_weight"),
                            "pallet_count": row.get("pallet_count"),
                        }
                    )
        finally:
            cursor.close()

        if not sp_rows:
            results.append(
                {
                    "site": site_code,
                    "year_month": month_label,
                    "weight_to_customer": 0.0,
                    "weight_to_consignment": 0.0,
                    "pallet_to_customer": 0.0,
                    "pallet_to_consignment": 0.0,
                }
            )
            continue

        df = pd.DataFrame(sp_rows)
        if "Ship_to_Customer" not in df:
            df["Ship_to_Customer"] = ""
        if "pick_weight" not in df:
            df["pick_weight"] = 0
        if "pallet_count" not in df:
            df["pallet_count"] = 0

        df["pick_weight"] = pd.to_numeric(df["pick_weight"], errors="coerce").fillna(0.0)
        df["pallet_count"] = pd.to_numeric(df["pallet_count"], errors="coerce").fillna(0.0)
        df["_customer_norm"] = df["Ship_to_Customer"].map(_normalize_customer_name)

        consignment_set = {
            _normalize_customer_name(name)
            for name in _SITE_CONSIGNMENT_CUSTOMERS.get(site_code, set())
        }
        consignment_mask = df["_customer_norm"].isin(consignment_set)

        weight_to_consignment = float(df.loc[consignment_mask, "pick_weight"].sum())
        pallet_to_consignment = float(df.loc[consignment_mask, "pallet_count"].sum())
        weight_total = float(df["pick_weight"].sum())
        pallet_total = float(df["pallet_count"].sum())

        results.append(
            {
                "site": site_code,
                "year_month": month_label,
                "weight_to_customer": weight_total - weight_to_consignment,
                "weight_to_consignment": weight_to_consignment,
                "pallet_to_customer": pallet_total - pallet_to_consignment,
                "pallet_to_consignment": pallet_to_consignment,
            }
        )

    return results


def _build_today_rows(
    conn,
    product_group: str,
    year_month: str | None = None,
) -> list[dict[str, object]]:
    """Build selected-month target-day customer-vs-consignment totals from SP output using pandas."""
    _month_start, _month_end, target_date, _month_label = _resolve_month_window(year_month)
    date_label = target_date.strftime("%Y/%m/%d")

    results: list[dict[str, object]] = []
    dbapi_conn = conn.connection

    for site_code in _MTD_SITES:
        cursor = dbapi_conn.cursor(dictionary=True)
        try:
            cursor.callproc(
                "sp_bl_lbs_cnt_carrier_customer",
                [
                    target_date.isoformat(),
                    target_date.isoformat(),
                    site_code,
                    product_group,
                ],
            )

            sp_rows: list[dict[str, object]] = []
            for result_set in cursor.stored_results():
                for row in result_set.fetchall():
                    sp_rows.append(
                        {
                            "Ship_to_Customer": row.get("Ship_to_Customer"),
                            "pick_weight": row.get("pick_weight"),
                            "pallet_count": row.get("pallet_count"),
                        }
                    )
        finally:
            cursor.close()

        if not sp_rows:
            results.append(
                {
                    "site": site_code,
                    "year_month": date_label,
                    "weight_to_customer": 0.0,
                    "weight_to_consignment": 0.0,
                    "pallet_to_customer": 0.0,
                    "pallet_to_consignment": 0.0,
                }
            )
            continue

        df = pd.DataFrame(sp_rows)
        if "Ship_to_Customer" not in df:
            df["Ship_to_Customer"] = ""
        if "pick_weight" not in df:
            df["pick_weight"] = 0
        if "pallet_count" not in df:
            df["pallet_count"] = 0

        df["pick_weight"] = pd.to_numeric(df["pick_weight"], errors="coerce").fillna(0.0)
        df["pallet_count"] = pd.to_numeric(df["pallet_count"], errors="coerce").fillna(0.0)
        df["_customer_norm"] = df["Ship_to_Customer"].map(_normalize_customer_name)

        consignment_set = {
            _normalize_customer_name(name)
            for name in _SITE_CONSIGNMENT_CUSTOMERS.get(site_code, set())
        }
        consignment_mask = df["_customer_norm"].isin(consignment_set)

        weight_to_consignment = float(df.loc[consignment_mask, "pick_weight"].sum())
        pallet_to_consignment = float(df.loc[consignment_mask, "pallet_count"].sum())
        weight_total = float(df["pick_weight"].sum())
        pallet_total = float(df["pallet_count"].sum())

        results.append(
            {
                "site": site_code,
                "year_month": date_label,
                "weight_to_customer": weight_total - weight_to_consignment,
                "weight_to_consignment": weight_to_consignment,
                "pallet_to_customer": pallet_total - pallet_to_consignment,
                "pallet_to_consignment": pallet_to_consignment,
            }
        )

    return results


@router.get("/weather/maxt1", include_in_schema=False)
async def weather_maxt1() -> FileResponse:
    """Serve the MaxT1 CONUS image with no-cache headers so the browser always fetches fresh."""
    return FileResponse(_ASSETS_DIR / "MaxT1_conus.png", headers=_NO_CACHE)


@router.get("/weather/national", include_in_schema=False)
async def weather_national() -> FileResponse:
    """Serve the national forecast image with no-cache headers."""
    return FileResponse(_ASSETS_DIR / "national_forecast.jpg", headers=_NO_CACHE)


@router.get(
    "/api/gas-prices",
    summary="Latest gas prices",
    description="Return the most recent national average gas prices scraped from AAA.",
)
async def gas_prices() -> JSONResponse:
    """Read the two most recent scrape dates and return today vs previous for comparison."""
    with _engine.connect() as conn:
        # Get the two most recent distinct scrape timestamps (date-level)
        dates_result = conn.execute(
            text("SELECT DISTINCT DATE(scraped_at) AS d FROM gas_prices ORDER BY d DESC LIMIT 2")
        )
        dates = [r.d for r in dates_result]

        if not dates:
            return JSONResponse([])

        latest_date = dates[0]
        prev_date = dates[1] if len(dates) > 1 else None

        # Fetch latest prices
        latest_result = conn.execute(
            text(
                "SELECT fuel_type, price, scraped_at FROM gas_prices"
                " WHERE DATE(scraped_at) = :d ORDER BY id DESC"
            ),
            {"d": latest_date},
        )
        # Keep only the most recent row per fuel type (in case of multiple runs same day)
        latest = {}
        scraped_at = None
        for r in latest_result:
            if r.fuel_type not in latest:
                latest[r.fuel_type] = float(r.price)
                if scraped_at is None:
                    scraped_at = r.scraped_at.strftime("%Y-%m-%d %H:%M:%S")

        # Fetch previous day prices for comparison
        prev = {}
        if prev_date:
            prev_result = conn.execute(
                text(
                    "SELECT fuel_type, price FROM gas_prices"
                    " WHERE DATE(scraped_at) = :d ORDER BY id DESC"
                ),
                {"d": prev_date},
            )
            for r in prev_result:
                if r.fuel_type not in prev:
                    prev[r.fuel_type] = float(r.price)

    rows = [
        {
            "fuel_type": fuel,
            "price": latest[fuel],
            "prev_price": prev.get(fuel),
            "scraped_at": scraped_at,
        }
        for fuel in ["Regular", "Mid-Grade", "Premium", "Diesel", "E85"]
        if fuel in latest
    ]
    return JSONResponse(rows)


@router.get(
    "/api/gas-prices/history",
    summary="Gas price history",
    description=(
        "Return historical national gas prices for all fuel types from the gas_prices table. "
        "Designed for time-series chart rendering on the Home page."
    ),
)
async def gas_prices_history() -> JSONResponse:
    """Read full gas price history ordered by scrape timestamp for line-chart visualization."""
    with _engine.connect() as conn:
        result = conn.execute(
            text(
                "SELECT id, fuel_type, price, scraped_at FROM gas_prices "
                "ORDER BY scraped_at ASC, id ASC"
            )
        )
        rows = [
            {
                "id": int(r.id),
                "fuel_type": r.fuel_type,
                "price": float(r.price),
                "scraped_at": r.scraped_at.strftime("%Y-%m-%d %H:%M:%S"),
            }
            for r in result
        ]
    return JSONResponse(rows)


@router.get(
    "/",
    response_class=HTMLResponse,
    summary="Home page",
    description="Main landing page displaying weather forecast images side by side.",
)
async def home(request: Request) -> HTMLResponse:
    """Render the home page."""
    today = datetime.date.today().strftime("%Y%m%d")
    return templates.TemplateResponse(
        "home/index.html",
        {"request": request, "active_page": "home", "cache_bust": today},
    )


@router.get(
    "/meeting-report",
    response_class=HTMLResponse,
    summary="Meeting Report page",
    description="Meeting reports and minutes sub-page under Home.",
)
async def meeting_report(request: Request) -> HTMLResponse:
    """Render the meeting report page with the filter form."""
    return templates.TemplateResponse(
        "home/meeting_report.html",
        {"request": request, "active_page": "meeting_report"},
    )


@router.get(
    "/api/meeting-report/all-site-mtd",
    response_class=HTMLResponse,
    summary="Meeting Report All Site MTD",
    description=(
        "Returns the All Site Month-to-Date shipped weight and pallet summary partial. "
        "This card is independent from Truck Appointment Date filter and uses selected year_month month-to-date."
    ),
)
async def meeting_report_all_site_mtd(
    request: Request,
    product_group: str = Query(..., description="Product group identifier"),
    year_month: str | None = Query(default=None, description="Year-month selector in YYYY-MM format"),
) -> HTMLResponse:
    """Render All Site MTD card partial independent of report date filter."""
    all_site_mtd_rows: list[dict[str, object]] = []
    error = None

    try:
        with _engine.connect() as conn:
            all_site_mtd_rows = _build_all_site_mtd_rows(
                conn=conn,
                product_group=product_group,
                year_month=year_month,
            )
    except Exception as exc:
        error = str(exc)

    return templates.TemplateResponse(
        "home/all_site_mtd_card.html",
        {
            "request": request,
            "all_site_mtd_rows": all_site_mtd_rows,
            "error": error,
            "product_group": product_group,
        },
    )


@router.get(
    "/api/meeting-report/today-summary",
    response_class=HTMLResponse,
    summary="Meeting Report Today Summary",
    description=(
        "Returns selected-month target-day shipped weight and pallet summary partial across all sites. "
        "Current month uses today; past months use month-end day."
    ),
)
async def meeting_report_today_summary(
    request: Request,
    product_group: str = Query(..., description="Product group identifier"),
    year_month: str | None = Query(default=None, description="Year-month selector in YYYY-MM format"),
) -> HTMLResponse:
    """Render Today Summary card partial."""
    today_rows: list[dict[str, object]] = []
    error = None

    try:
        with _engine.connect() as conn:
            today_rows = _build_today_rows(
                conn=conn,
                product_group=product_group,
                year_month=year_month,
            )
    except Exception as exc:
        error = str(exc)

    return templates.TemplateResponse(
        "home/today_summary_card.html",
        {
            "request": request,
            "today_rows": today_rows,
            "error": error,
            "product_group": product_group,
        },
    )


@router.get(
    "/api/meeting-report/warehouse-pallet-movement-mtd",
    response_class=HTMLResponse,
    summary="Meeting Report Warehouse Pallet Movement MTD",
    description=(
        "Returns standalone Warehouse Pallet Movement month-to-date summary "
        "for AMJK SW using selected year_month month-to-date window."
    ),
)
async def meeting_report_warehouse_pallet_movement_mtd(request: Request) -> HTMLResponse:
    """Render Warehouse Pallet Movement MTD card partial independent of report filters."""
    year_month = request.query_params.get("year_month")
    today = datetime.date.today()
    month_label = today.strftime("%Y/%m")
    date_from = today.replace(day=1).isoformat()
    date_to = today.isoformat()
    entry_to_date = 0
    exit_to_date = 0
    shipped_to_date = 0
    error = None

    try:
        month_start, _month_end, effective_end, month_label = _resolve_month_window(year_month)
        date_from = month_start.isoformat()
        date_to = effective_end.isoformat()

        with _engine.connect() as conn:
            entry_exit_row = conn.execute(
                _PALLET_ENTRY_EXIT_SQL,
                {"date_from": date_from, "date_to": date_to},
            ).fetchone()
            entry_to_date = round(float(entry_exit_row.entry_to_date) if entry_exit_row and entry_exit_row.entry_to_date else 0)
            exit_to_date = round(float(entry_exit_row.exit_to_date) if entry_exit_row and entry_exit_row.exit_to_date else 0)

            shipped_to_date = int(conn.execute(text("""
                SELECT SUM(pallet_count)
                FROM warship.vw_bl_lbs_cnt_carrier_customer
                WHERE site = 'AMJK'
                  AND product_group = 'SW'
                  AND Truck_Appointment_Date BETWEEN :date_from AND :date_to
            """), {"date_from": date_from, "date_to": date_to}).scalar() or 0)
    except Exception as exc:
        error = str(exc)

    return templates.TemplateResponse(
        "home/warehouse_pallet_movement_mtd_card.html",
        {
            "request": request,
            "month_label": month_label,
            "entry_to_date": entry_to_date,
            "exit_to_date": exit_to_date,
            "shipped_to_date": shipped_to_date,
            "error": error,
        },
    )


@router.get(
    "/api/meeting-report/results",
    response_class=HTMLResponse,
    summary="Meeting Report results",
    description=(
        "Executes the meeting report query filtered by site, product_group, and date. "
        "Returns an HTML partial (cards) consumed by HTMX on the Meeting Report page."
    ),
)
async def meeting_report_results(
    request: Request,
    site: str = Query(..., description="Site code, e.g. 'CFP'"),
    product_group: str = Query(..., description="Product group identifier"),
    date: str = Query(..., description="Truck appointment date (YYYY-MM-DD)"),
) -> HTMLResponse:
    """
    Run the aggregated meeting report query and return rendered result cards.
    On DB error, the partial displays an error alert instead of crashing the page.
    """
    rows = []
    consignment_count = 0
    custom_count = 0
    error = None

    try:
        selected_date = datetime.date.fromisoformat(date.strip())
    except ValueError as exc:
        error = f"Invalid date: {exc}"
        selected_date = None

    try:
        if selected_date is None:
            raise ValueError(error)

        with _engine.connect() as conn:
            result = conn.execute(
                _MEETING_REPORT_SQL,
                {
                    "site": site,
                    "product_group": product_group,
                    "date": selected_date.isoformat(),
                },
            )
            # Convert each row to a plain dict, casting Decimal → float so
            # Jinja2 arithmetic and Python number formatting work correctly.
            # MySQL SUM/ROUND returns decimal.Decimal which is not JSON/template-safe.
            numeric_keys = {"Pallets", "Weight", "Freight", "Avg_Freight_per_lb"}
            rows = [
                {
                    k: float(v) if k in numeric_keys and v is not None else v
                    for k, v in row._mapping.items()
                }
                for row in result.fetchall()
            ]

            params = {
                "site": site,
                "product_group": product_group,
                "date": selected_date.isoformat(),
            }
            consignment_count = int(conn.execute(_CONSIGNMENT_COUNT_SQL, params).scalar() or 0)
            custom_count = int(conn.execute(_CUSTOM_COUNT_SQL, params).scalar() or 0)

    except Exception as exc:
        error = str(exc)

    return templates.TemplateResponse(
        "home/meeting_report_results.html",
        {
            "request": request,
            "rows": rows,
            "consignment_count": consignment_count,
            "custom_count": custom_count,
            "error": error,
            "site": site,
            "product_group": product_group,
            "date": date,
        },
    )


@router.get(
    "/briefing",
    response_class=HTMLResponse,
    summary="VIP Operations Briefing page",
    description="Printable snapshot of warehouse and shipping operations for VIP visits.",
)
async def briefing(request: Request) -> HTMLResponse:
    """Render the VIP Operations Briefing page."""
    return templates.TemplateResponse(
        "home/briefing.html",
        {"request": request, "active_page": "briefing"},
    )


# ---------------------------------------------------------------------------
# Analytics — weight by year
# ---------------------------------------------------------------------------
# this is from mei excel book
_WEIGHT_BY_YEAR_SQL = text("""
    SELECT
        YEAR(Truck_Appointment_Date)   AS year,
        MONTH(Truck_Appointment_Date)  AS month,
        SUM(pick_weight)               AS total_weight
    FROM warship.vw_bl_lbs_cnt_carrier_customer
    WHERE Site          = :site
      AND Product_Group = :product_group
    GROUP BY YEAR(Truck_Appointment_Date), MONTH(Truck_Appointment_Date)
    ORDER BY year, month
""")


@router.get(
    "/api/analytics/weight-by-year",
    summary="Daily pick_weight split by year",
    description=(
        "Returns monthly pick_weight totals grouped into per-year series. "
        "Each element in the array is one year: {year, data:[{month, weight}]}. "
        "Suitable for rendering as a multi-line Plotly chart — one line per year."
    ),
)
async def weight_by_year(
    site: str = Query(default="AMJK", description="Site code, e.g. 'AMJK'"),
    product_group: str = Query(default="SW", description="Product group, e.g. 'SW'"),
) -> JSONResponse:
    """
    Query pick_weight by Truck_Appointment_Date and group results into
    per-year series — one series object per calendar year found in the data.
    """
    try:
        with _engine.connect() as conn:
            rows = conn.execute(
                _WEIGHT_BY_YEAR_SQL,
                {"site": site, "product_group": product_group},
            ).fetchall()
    except Exception as exc:
        return JSONResponse(status_code=500, content={"error": str(exc)})

    # Group rows by year; each year becomes one line on the chart
    by_year: dict[int, list] = defaultdict(list)
    for row in rows:
        by_year[int(row.year)].append({
            "month":  int(row.month),
            "weight": float(row.total_weight or 0),
        })

    result = [
        {"year": year, "data": data}
        for year, data in sorted(by_year.items())
    ]
    return JSONResponse(content=result)


# ---------------------------------------------------------------------------
# Analytics — freight lbs by year (frt_cost_breakdown_mei)
# ---------------------------------------------------------------------------

_FREIGHT_LBS_BY_YEAR_SQL = text("""
    SELECT
        yyyy          AS year,
        mm            AS month,
        SUM(lbs)      AS total_lbs
    FROM warship.frt_cost_breakdown_mei
    WHERE site = :site
    GROUP BY yyyy, mm
    ORDER BY yyyy, mm
""")


@router.get(
    "/api/analytics/freight-lbs-by-year-mei",
    summary="Monthly freight lbs by year from frt_cost_breakdown_mei",
    description=(
        "Returns monthly lbs totals from frt_cost_breakdown_mei grouped into "
        "per-year series. Each element: {year, data:[{month, lbs}]}. "
        "Suitable for a multi-line Plotly chart — one line per year. "
        "param site defaults to 'SW'."
    ),
)
async def freight_lbs_by_year(
    site: str = Query(default="SW", description="Site/product-group code, e.g. 'SW'"),
) -> JSONResponse:
    """
    Query lbs by yyyy/mm from frt_cost_breakdown_mei filtered by site,
    and group results into per-year series — one series object per year.
    """
    try:
        with _engine.connect() as conn:
            rows = conn.execute(
                _FREIGHT_LBS_BY_YEAR_SQL,
                {"site": site},
            ).fetchall()
    except Exception as exc:
        return JSONResponse(status_code=500, content={"error": str(exc)})

    by_year: dict[int, list] = defaultdict(list)
    for row in rows:
        by_year[int(row.year)].append({
            "month": int(row.month),
            "lbs":   float(row.total_lbs or 0),
        })

    result = [
        {"year": year, "data": data}
        for year, data in sorted(by_year.items())
    ]
    return JSONResponse(content=result)


# ---------------------------------------------------------------------------
# Analytics — unit freight cost (unit_frt_cost_john)
# ---------------------------------------------------------------------------

_UNIT_FRT_COST_JOHN_SQL = text("""
    SELECT
        id,
        yyyy,
        mm,
        division,
        product,
        wt_lbs,
        freight
    FROM warship.unit_frt_cost_john
""")


@router.get(
    "/api/analytics/unit-frt-cost-john",
    summary="All rows from unit_frt_cost_john",
    description=(
        "Returns every row from warship.unit_frt_cost_john as a JSON array. "
        "Each element: {id, yyyy, mm, division, product, wt_lbs, freight}."
    ),
)
async def unit_frt_cost_john() -> JSONResponse:
    """Fetch all rows from unit_frt_cost_john and return as a JSON array."""
    try:
        with _engine.connect() as conn:
            rows = conn.execute(_UNIT_FRT_COST_JOHN_SQL).fetchall()
    except Exception as exc:
        return JSONResponse(status_code=500, content={"error": str(exc)})

    result = [
        {
            "id":       row.id,
            "yyyy":     row.yyyy,
            "mm":       row.mm,
            "division": row.division,
            "product":  row.product,
            "wt_lbs":   float(row.wt_lbs)  if row.wt_lbs  is not None else None,
            "freight":  float(row.freight) if row.freight is not None else None,
        }
        for row in rows
    ]
    return JSONResponse(content=result)


# ---------------------------------------------------------------------------
# Analytics — annual freight cost by plant (Excel workbook)
# ---------------------------------------------------------------------------

_EXCEL_FRTCOST_PATH = (
    Path(__file__).parent.parent
    / "raw_data" / "Mei" / "AMJK Frt cost breakdown by plants.xlsx"
)
_PLANT_NAMES = ("BP", "SW", "CT", "YA", "Total")


@router.get(
    "/api/analytics/freight-cost-by-plant",
    summary="Annual freight cost by plant from Excel workbook",
    description=(
        "Reads the 'AMJK Frt cost breakdown by plants' Excel workbook and returns "
        "annual YTD freight cost ($) per plant (BP, SW, CT, YA, Total) for 2019–2026. "
        "2018 is excluded because that sheet tracks weight (lbs), not cost ($)."
    ),
)
async def freight_cost_by_plant() -> JSONResponse:
    """Parse the AMJK freight cost Excel workbook and return annual YTD totals by plant."""
    try:
        wb = openpyxl.load_workbook(_EXCEL_FRTCOST_PATH, data_only=True, read_only=True)
    except Exception as exc:
        return JSONResponse(status_code=500, content={"error": str(exc)})

    result = []
    for year_str in [str(y) for y in range(2019, 2027)]:
        if year_str not in wb.sheetnames:
            continue
        ws = wb[year_str]
        # Data rows start at row 4 (1-indexed): BP, SW, CT, YA, Total
        # YTD Total is the second-to-last column; last column is % share
        year_data: dict = {"year": int(year_str)}
        for row in ws.iter_rows(min_row=4, max_row=8, values_only=True):
            if not row or row[0] not in _PLANT_NAMES:
                continue
            ytd = row[-2]
            year_data[row[0]] = float(ytd) if isinstance(ytd, (int, float)) else 0.0
        result.append(year_data)

    wb.close()
    return JSONResponse(content=result)


# ---------------------------------------------------------------------------
# Analytics — Product Trend (from Warship DB via stored procedure)
# ---------------------------------------------------------------------------

@router.get(
    "/api/analytics/product-trend-top",
    summary="Top N products by total weight from stored procedure data",
    description=(
        "Loads shipped-product rows from sp_get_all_shipped_product and returns the top N "
        "products ranked by total shipped weight. "
        "Each product includes: product_code, total_weight, shipment_count, avg_weight. "
        "Useful for identifying portfolio leaders and consolidation trends."
    ),
)
async def product_trend_top(
    top_n: int = Query(default=10, ge=1, le=50, description="Number of top products to return")
) -> JSONResponse:
    """Get top N products by total weight."""
    try:
        rows = load_product_data_from_sp()
        products = get_top_products(rows, top_n=top_n)
        return JSONResponse(content=products)
    except Exception as exc:
        return JSONResponse(status_code=500, content={"error": str(exc)})


@router.get(
    "/api/analytics/product-trend-monthly",
    summary="Monthly trend data for top products",
    description=(
        "Returns monthly totals (weight, shipment count) for each of the top N products. "
        "Useful for multi-line chart showing growth trajectories. "
        "Each element: {product_code, monthly: [{year_month, total_weight, shipment_count, avg_weight}]}"
    ),
)
async def product_trend_monthly(
        top_n: int = Query(default=5, ge=1, le=20, description="Number of top products")
    ) -> JSONResponse:
        """Get monthly trend data for top products."""
        try:
            rows = load_product_data_from_sp()
            products = get_top_products(rows, top_n=top_n)
    
            result = []
            for product in products:
                monthly = get_monthly_trend(rows, product["product_code"])
                result.append({
                    "product_code": product["product_code"],
                    "total_weight": product["total_weight"],
                    "shipment_count": product["shipment_count"],
                    "monthly": monthly
                })
    
            return JSONResponse(content=result)
        except Exception as exc:
            return JSONResponse(status_code=500, content={"error": str(exc)})


@router.get(
    "/api/analytics/product-diversity",
    summary="Product portfolio diversity by month",
    description=(
        "Returns unique product count per month, showing portfolio consolidation trend. "
        "Each element: {year_month, unique_products, total_weight, total_shipments}"
    ),
)
async def product_diversity() -> JSONResponse:
    """Get monthly product diversity metrics."""
    try:
        rows = load_product_data_from_sp()
        diversity = get_product_diversity_over_time(rows)
        return JSONResponse(content=diversity)
    except Exception as exc:
        return JSONResponse(status_code=500, content={"error": str(exc)})


# ---------------------------------------------------------------------------
# Analytics — SW transport type by year (Transp Type.xlsx, sheet AM)
# ---------------------------------------------------------------------------

_EXCEL_TRANSP_PATH = (
    Path(__file__).parent.parent
    / "raw_data" / "John" / "Transp Type.xlsx"
)

# Row order within each product block (11 rows per product)
_TRANSP_BLOCK_SIZE = 11
_TRANSP_TYPES = [
    "Intermodal", "FTL", "Railcar", "LTL", "Export",
    "Sample", "Reconsignment", "Prepaid Total", "CPU (Lolita)",
    "Wharehouse", "Grand Total",
]
# SW is the second product block (0-indexed: 1)
_SW_BLOCK_INDEX = 1


@router.get(
    "/api/analytics/sw-transport-type-by-year",
    summary="SW annual shipping volume by transport type (AMTOPP)",
    description=(
        "Reads 'Transp Type.xlsx' (AM sheet) and returns annual totals "
        "(lbs) for SW product by transport type: FTL, LTL, Intermodal, Export, "
        "Prepaid Total. Covers 2020–2025; 2026 is excluded as a partial year."
    ),
)
async def sw_transport_type_by_year() -> JSONResponse:
    """Parse the AM sheet and return SW annual lbs totals per transport type."""
    try:
        wb = openpyxl.load_workbook(_EXCEL_TRANSP_PATH, data_only=True, read_only=True)
    except Exception as exc:
        return JSONResponse(status_code=500, content={"error": str(exc)})

    ws = wb["AM"]
    all_rows = list(ws.iter_rows(min_row=4, max_row=55, values_only=True))
    header = all_rows[0]

    # Period columns are even-indexed starting at col 2; label format "YY/MM"
    periods = [
        (ci, str(header[ci]).strip())
        for ci in range(2, len(header), 2)
        if header[ci] is not None and "/" in str(header[ci])
    ]

    # Extract the SW block
    block_start = 2 + _SW_BLOCK_INDEX * _TRANSP_BLOCK_SIZE
    block_rows = all_rows[block_start: block_start + _TRANSP_BLOCK_SIZE]

    # Accumulate annual totals for the types we care about
    keep = {"FTL", "LTL", "Intermodal", "Export", "Prepaid Total"}
    year_totals: dict[int, dict[str, float]] = {}

    for t_idx, ttype in enumerate(_TRANSP_TYPES):
        if ttype not in keep:
            continue
        row = block_rows[t_idx]
        for ci, period in periods:
            yy, _ = period.split("/")
            year = int("20" + yy.strip())
            if year > 2025:          # exclude partial 2026
                continue
            val = row[ci] if ci < len(row) and isinstance(row[ci], (int, float)) else 0.0
            year_totals.setdefault(year, {t: 0.0 for t in keep})
            year_totals[year][ttype] += val

    wb.close()

    result = [
        {
            "year":          year,
            "FTL":           round(d["FTL"], 0),
            "LTL":           round(d["LTL"], 0),
            "Intermodal":    round(d["Intermodal"], 0),
            "Export":        round(d["Export"], 0),
            "Prepaid_Total": round(d["Prepaid Total"], 0),
        }
        for year, d in sorted(year_totals.items())
    ]
    return JSONResponse(content=result)



# ── AMJK Freight Monthly Avg Comparison ──────────────────────────────────────
# Reads SW row from _EXCEL_FRTCOST_PATH (same file as freight-cost-by-plant).
# Historical = full calendar years 2018–(cur_year-1); YTD = current year sheet.
# NOTE: 2018 sheet layout is reversed — lbs section is first, Frt Amt is second.
# All other years: Frt Amt section is first (rows 3–8), lbs section second (10–15).


@router.get(
    "/api/analytics/amjk-frt-ytd-vs-avg",
    summary="AMJK SW freight cost monthly avg comparison",
    description=(
        "Reads the SW row from 'AMJK Frt cost breakdown by plants.xlsx'. "
        "Historical avg = monthly avg across full calendar years (2018–last year). "
        "YTD = current year sheet totals. Returns Frt Amt ($), lbs, and ¢/lb."
    ),
)
async def amjk_frt_ytd_vs_avg() -> JSONResponse:
    """Return AMJK SW monthly avg (historical full years 2018+) and current YTD from Excel."""
    import datetime as _dt

    cur_year = _dt.date.today().year

    try:
        wb = openpyxl.load_workbook(_EXCEL_FRTCOST_PATH, data_only=True, read_only=True)
    except Exception as exc:
        return JSONResponse(status_code=500, content={"error": str(exc)})

    hist_frt_total = 0.0
    hist_wt_total  = 0.0
    hist_months    = 0
    hist_years: list[int] = []
    ytd_frt = 0.0
    ytd_wt  = 0.0
    ytd_n   = 0

    for year_str in [str(y) for y in range(2018, cur_year + 1)]:
        if year_str not in wb.sheetnames:
            continue
        ws = wb[year_str]
        rows = list(ws.iter_rows(min_row=1, max_row=15, values_only=True))

        # 2018 sheet is reversed: top section = lbs, bottom section = Frt Amt ($)
        # All other years: top section = Frt Amt ($), bottom section = lbs
        top_is_lbs = (rows[2] and str(rows[2][0]).startswith("Weight"))
        if top_is_lbs:
            sw_frt_row = next((r for r in rows[9:15] if r and r[0] == "SW"), None)
            sw_lbs_row = next((r for r in rows[2:8]  if r and r[0] == "SW"), None)
        else:
            sw_frt_row = next((r for r in rows[2:8]  if r and r[0] == "SW"), None)
            sw_lbs_row = next((r for r in rows[9:15] if r and r[0] == "SW"), None)

        if sw_frt_row is None or sw_lbs_row is None:
            continue

        ytd_frt_val = sw_frt_row[-2]
        ytd_wt_val  = sw_lbs_row[-2]

        if not isinstance(ytd_frt_val, (int, float)) or not isinstance(ytd_wt_val, (int, float)):
            continue

        yr = int(year_str)
        # Monthly data sits at even-indexed columns 2,4,6,...,24 (skipping Vari columns)
        month_count = sum(
            1 for i in range(2, 26, 2)
            if i < len(sw_frt_row) and isinstance(sw_frt_row[i], (int, float))
        )

        if yr < cur_year:
            if month_count == 12:  # only include complete years in historical avg
                hist_frt_total += float(ytd_frt_val)
                hist_wt_total  += float(ytd_wt_val)
                hist_months    += 12
                hist_years.append(yr)
        else:
            ytd_frt = float(ytd_frt_val)
            ytd_wt  = float(ytd_wt_val)
            ytd_n   = month_count

    wb.close()

    if not hist_years:
        return JSONResponse(status_code=500, content={"error": "No historical data found"})

    avg_monthly_frt = hist_frt_total / hist_months
    avg_monthly_wt  = hist_wt_total  / hist_months
    avg_cperlb  = (avg_monthly_frt / avg_monthly_wt * 100) if avg_monthly_wt else 0.0
    ytd_cperlb  = (ytd_frt / ytd_wt * 100) if ytd_wt else 0.0

    return JSONResponse(content={
        "avg": {
            "frt":    round(avg_monthly_frt, 0),
            "wt":     round(avg_monthly_wt, 0),
            "cperlb": round(avg_cperlb, 4),
            "label":  f"Monthly Avg ({hist_years[0]}–{hist_years[-1]})",
        },
        "ytd": {
            "frt":    round(ytd_frt, 0),
            "wt":     round(ytd_wt, 0),
            "cperlb": round(ytd_cperlb, 4),
            "months": ytd_n,
            "label":  f"YTD {cur_year} ({ytd_n} mo.)",
        },
    })
